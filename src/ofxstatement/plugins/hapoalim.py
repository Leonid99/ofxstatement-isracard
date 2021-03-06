#!/usr/bin/env python3
import sys
from bs4 import BeautifulSoup
import dateutil.parser
import openpyxl
import zipfile

from ofxstatement.plugin import Plugin
from ofxstatement.parser import StatementParser
from ofxstatement.statement import StatementLine, Statement
from ofxstatement.exceptions import ParseError


class HapoalimPlugin(Plugin):
    """Hapoalim plugin (for developers only)
    """

    def get_parser(self, filename):
        return HapoalimParser(filename)

class HapoalimParser(StatementParser):
    def __init__(self, filename):
        self.filename = filename

    def log(self, msg):
        print(msg, file=sys.stderr)

    def detect_version(self):
        def validator_old(filename):
            try:
                with open(filename, 'r', encoding='iso-8859-8') as f:
                    bs = BeautifulSoup(f, 'lxml')
                table = bs.find_all('table', id='mytable_body')[0]
                return table.find_all('th', id='header1')[0].string == "תאריך"
            except (UnicodeDecodeError, IndexError):
                return False
            return False

        def validator_new(filename):
            try:
                with open(filename, 'r', encoding='iso-8859-8') as f:
                    bs = BeautifulSoup(f, 'lxml')
                table = bs.find_all('table', i__d='trBlueOnWhite12')[0]
                return table.tr.td.string == "תאריך"
            except (UnicodeDecodeError, IndexError):
                return False
            return False

        def validator_xslx(filename):
            try:
                with open(filename, 'rb') as f:
                    wb = openpyxl.load_workbook(f)
                if wb.sheetnames[0] != 'גיליון1':
                    return False
                return wb.worksheets[0]['A6'].value == "תאריך"
            except (openpyxl.utils.exceptions.InvalidFileException, zipfile.BadZipFile):
                return False
            return False

        VALIDATORS = [validator_old, validator_new, validator_xslx]

        for n, validator in enumerate(VALIDATORS):
            if validator(self.filename):
                return n
        return None

    @staticmethod
    def get_float(s):
        NBSP = u'\xa0'
        if s in ['', NBSP]:
            return None
        return float(s.replace(',', ''))

    def parse(self):
        DATA_DATE = 'data_date'
        DATA_AMT = 'data_amt'
        DATA_DSC = 'data_dsc'
        DATA_MEMO = 'data_mem'
        DATA_BAL = 'data_bal'

        def parser_old(filename):

            def parse_tr(tr):
                def get_header(tr, n):
                    return str(tr.find_all(headers='header'+str(n))[0].string)

                hdrs = tr.find_all(headers=True)
                assert len(hdrs) in [0, 7]
                if len(hdrs) == 0:
                    cont = True
                    date_or_comm = str(tr.find_all(colspan='5')[0].string)
                    desc = None
                    paym = None
                    depo = None
                    bal = None
                else:
                    cont = False
                    date_or_comm = dateutil.parser.parse(get_header(tr, 1), dayfirst=True)
                    desc = get_header(tr, 2)

                    paym = self.get_float(get_header(tr, 5))
                    depo = self.get_float(get_header(tr, 6))

                    assert (depo is None) != (paym is None), 'Either depo or paym need to exist, but not both: ' + str(tr)

                    bal = self.get_float(get_header(tr, 7))

                return cont, date_or_comm, desc, paym, depo, bal

            with open(filename, 'r', encoding='iso-8859-8') as f:
                soup = BeautifulSoup(f, 'lxml')
                statement = Statement(currency='ILS')
                data = []
                for tr in soup.find_all('table', id='mytable_body')[0].find_all(id='TR_ROW_BANKTABLE'):
                    cont, date_or_comm, desc, paym, depo, bal = parse_tr(tr)
                    if cont:
                        data[-1][DATA_MEMO] = date_or_comm
                    else:
                        new_line = {}

                        new_line[DATA_DATE] = date_or_comm
                        new_line[DATA_DSC] = desc
                        new_line[DATA_AMT] = -paym if paym is not None else depo
                        new_line[DATA_BAL] = bal
                        new_line[DATA_MEMO] = None

                        data.append(new_line)
                return data

        def parser_new(filename):
            with open(filename, 'r', encoding='iso-8859-8') as f:
                bs = BeautifulSoup(f, 'lxml')
            trs = bs.find_all('table', i__d='trBlueOnWhite12')[0].find_all('tr', recursive=False)

            data = []
            for tr in trs[1:]:
                new_line = {}
                tds = tr.find_all('td')
                if len(tds) != 7:
                    continue
                new_line[DATA_DATE] = dateutil.parser.parse(str(tds[0].string), dayfirst=True)
                new_line[DATA_DSC] = str(tds[1].string)

                paym = self.get_float(str(tds[4].string))
                depo = self.get_float(str(tds[5].string))
                assert (depo is None) != (paym is None), 'Either depo or paym need to exist, but not both: ' + str(tds)
                new_line[DATA_AMT] = -paym if paym is not None else depo

                new_line[DATA_BAL] = self.get_float(str(tds[6].string))

                new_line[DATA_MEMO] = None

                data.append(new_line)

            return data

        def parser_xslx(filename):
            with open(filename, 'rb') as f:
                ws = openpyxl.load_workbook(f).worksheets[0]

            iterrows = iter(ws.rows)
            for _ in range(6):
                next(iterrows)

            data = []
            for row in iterrows:
                new_line = {}
                new_line[DATA_DATE] = dateutil.parser.parse(str(row[0].value), dayfirst=False)
                new_line[DATA_DSC] = str(row[1].value)

                paym = self.get_float(str(row[3].value))
                depo = self.get_float(str(row[4].value))
                assert (depo is None) != (paym is None), 'Either depo or paym need to exist, but not both: '
                new_line[DATA_AMT] = -paym if paym is not None else depo

                new_line[DATA_BAL] = self.get_float(str(row[5].value))

                new_line[DATA_MEMO] = " ".join([str(c.value) for c in row[7:] if c.value is not None])

                data.append(new_line)

            return data
        PARSERS = [parser_old, parser_new, parser_xslx]

        v = self.detect_version()
        assert v is not None, "Unsupported file %s" % self.filename
        self.log("Detected file of version %d" % v)

        data = PARSERS[v](self.filename)
        self.log('Found %d transactions' % len(data))

        stmnt = Statement(currency='ILS')
        for d in data:
            stmt_line = StatementLine(date=d[DATA_DATE], amount=d[DATA_AMT])
            #stmt_line.end_balance = d[DATA_BAL] #TODO: conf
            stmt_line.payee = d[DATA_DSC]
            if d[DATA_MEMO] is not None:
                stmt_line.memo = d[DATA_MEMO]

            if stmt_line.payee.startswith('משיכה'):
                stmt_line.trntype = 'ATM'
            elif stmt_line.payee.startswith('שיק'):
                stmt_line.trntype = 'CHECK'
            else:
                stmt_line.trntype = "CASH" if d[DATA_AMT] < 0 else "DEP"

            stmt_line.assert_valid()
            stmnt.lines.append(stmt_line)
        return stmnt
